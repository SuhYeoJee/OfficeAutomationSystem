from PyQt5.QtWidgets import QMessageBox, QHBoxLayout, QVBoxLayout
from PyQt5.QtGui import QPainter, QImage
from mw import AppController
from ipoForm import ItemOrderForm
from PyQt5.QtCore import Qt
from igzg.utils import getNowStr
from math import ceil
from collections import defaultdict
from bt import BorderTable

class ItemOrder(ItemOrderForm):
    def __init__(self, windowTitle:str, ac:AppController, net=0):
        ItemOrderForm.__init__(self)
        self.ac = ac
        self.setWindowTitle(windowTitle)
        self.initUI()
        # self.bondSelected = False
        # self.segmentSelected = False
        # self.selectedBond = {}
        # self.selectedSegment = {}
        self.tableName = 'iporder'
        self.tablePK = 'mfgOrderNo'
        # self.net = net if net != 0 else "재고"
        self.voidDict = defaultdict(lambda: '')
        self.initVal()

        # print(self.getText(*self.tableWidgetRCDict["Name"])) # 테이블 값 조회 가능
    # ===========================================================================================
    # 화면 그리기
    def initUI(self):
        self.layout = QHBoxLayout()
        self.setLayout(self.layout)

        self.tableWidget = BorderTable()
        self.tableWidget.itemChanged.connect(self.onItemChanged)

        self.tableWidget.boxInfo = self.getBoxInfo()
        self.tableWidget.setWordWrap(False)
        self.tableWidget.setFixedSize(1300, 1650)
        self.resize(2000, 1650)
        self.layout.addWidget(self.tableWidget)    
        self.setNewForm()

        # groupSelect
        self.groupNames = list(set(self.ac.dm.getSingleColDatas('product','groupName')))
        self.setComboBox(*self.tableWidgetRCDict["productName1"],['제품선택']+self.groupNames, self.selectGroup)
        # -------------------------------------------------------------------------------------------
        # # segmentSelect

        # self.segmentNames = self.ac.dm.getSingleColDatas("segment","seg_no")
        # self.setComboBox(*self.tableWidgetRCDict["segNo1"],['세그먼트선택']+self.segmentNames, self.selectSegment)
        # # -------------------------------------------------------------------------------------------
        # # sporderSelect (0,0)

        # self.sporderNames = self.ac.dm.getSingleColDatas("sporder","mfgOrderNo")
        # self.setComboBox(0,0,['']+self.sporderNames, self.selectSporder)
        # # -------------------------------------------------------------------------------------------        
        # # 우측 입력창
        calLayout = QVBoxLayout()
        calLayout.addLayout(self.getButtonLayout())
        self.layout.addLayout(calLayout)


    # 초깃값 입력
    def initVal(self):
        # 날짜 가져오기
        self.setText(*self.tableWidgetRCDict["date"],getNowStr("%Y년 %m월 %d일"))


    # 내용 비우기
    def clearAll(self):
        for k in self.tableWidgetRCDict.keys():
            #table widget
            if k in ['bondSelect', 'segNo1']:
                self.getItem(*self.tableWidgetRCDict[k]).setCurrentIndex(0)
                self.setText(*self.tableWidgetRCDict[k],'')
            elif 'Check' in k:
                self.getItem(*self.tableWidgetRCDict[k]).setChecked(False)
            else:
                self.setText(*self.tableWidgetRCDict[k],'')
        self.initVal()


# ===========================================================================================
    # 표 수정시 호출
    def onItemChanged(self, item):
        pos = (item.row(), item.column())
        print(pos)

        try:
            if  self.groupSelected and pos in [self.tableWidgetRCDict['tab{}Amount'.format(x+1)] for x in range(4)] : # 탭 수량 입력
                self.amountUpdate()

        except: ...
    # -------------------------------------------------------------------------------------------
    # 저장버튼 클릭시 호출
    def saveTableImage(self):
        tableImage = QImage(self.tableWidget.viewport().size(), QImage.Format_ARGB32)
        tableImage.fill(Qt.transparent)

        painter = QPainter(tableImage)
        self.tableWidget.viewport().render(painter)
        painter.end()

        tableImage.save("tableImage.png")
        print("이미지 저장 완료.")        
    # -------------------------------------------------------------------------------------------
    # db삽입용 dict로 만들기
    def getDataDict(self):
        dataDict = {}

        # tableWidget에서
        for k, v in self.tableWidgetRCDict.items():
            dataDict[k] = str(self.getText(*v))
        else:
            dataDict[self.tablePK] = dataDict["mfgOrderNo1"].strip() + dataDict["mfgOrderNo2"].strip()
        
        dataDict = {x:dataDict[x] if dataDict[x] != 'None' else '' for x in dataDict}
        return dataDict
    
    # DB에 전체 저장
    def saveToDB(self):
        dataDict = self.getDataDict()
        try:
            # PK 확인
            if dataDict[self.tablePK] == '': raise
        except:
            QMessageBox.warning(self, "PK누락", f"필수항목이 입력되지 않았습니다. \n{self.tablePK}")
        else:
            if self.ac.dm.insertDataUpdateNone(self.tableName, dataDict) == False: #DB입력
                reply = QMessageBox.question(self, "PK중복", f"이미 존재하는 항목입니다. 덮어쓰시겠습니까? \n{self.tablePK}: {dataDict[self.tablePK]}",
                                                QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if reply == QMessageBox.Yes:
                    self.ac.dm.insertDataUpdateAll(self.tableName, dataDict) #DB입력
                else: return
    # -------------------------------------------------------------------------------------------
    # 엑셀로 저장
    def saveToElsx(self):

        spFilename = f'SP-ORDER-{getNowStr("%Y")}.xlsx'
        tempFilename = 'template.xlsx'
        sheetName = getNowStr("%Y") + '-' + self.getText(*self.tableWidgetRCDict["mfgOrderNo2"])

        import os
        from shutil import copyfile
        if not os.path.exists(spFilename):
            copyfile(tempFilename, spFilename)

        from openpyxl import load_workbook
        wb = load_workbook(spFilename)
        newSp = wb.copy_worksheet(wb['sptemp'])
        newSp.title = sheetName

        for v in self.tableWidgetRCDict.values():
            r,c = v
            newSp.cell(row=r+1, column=c+1,value=self.getText(*v))

        for v in self.configTableRCDict.values():
            r,c = v
            newSp.cell(row=r+20, column=c+12,value=self.getText(*v,self.configTable))

        for k,v in self.labelExcelRCDict.items():
            r,c = v
            try:
                res = str(float(self.labelDict[k].text().split(':')[-1].strip()))
            except:
                res = ''
            newSp.cell(row=r, column=c,value=res)

        wb.save(spFilename)
        QMessageBox.information(None, "저장완료", f"[{spFilename}] {sheetName} 저장완료", QMessageBox.Ok)


# ==========================================================================================
    # 그룹 선택시 호출
    def selectGroup(self,index):
        self.clearAll()
        if index == 0: 
            groupName = ''
            self.selectedGroupItems = []
            self.groupSelected = False
        else:
            groupName = self.groupNames[index-1]
            self.selectedGroupItems = self.ac.dm.getMultiItemDatas('product','groupName',groupName)

            self.items = {"tab":[],"segment":[],"shank":[],"sub":[]}
            self.loadGroupItem(groupName)


    def getTabName(self,groupName,name):
            return name if name.replace(groupName,'').strip() == '' else name.replace(groupName,'').strip()
    # -------------------------------------------------------------------------------------------
    # 그룹 아이템 정보 불러오기
    def loadGroupItem(self, groupName):
        for i in self.selectedGroupItems:
            self.items["tab"].append(self.getTabName(groupName,str(i["name"])))
            if i["seg1No"] and i["seg1No"] not in self.items["segment"]:self.items["segment"].append(str(i["seg1No"]).strip())
            if i["seg2No"] and i["seg2No"] not in self.items["segment"]:self.items["segment"].append(str(i["seg2No"]).strip())
            if i["shank"] and i["shank"] not in self.items["shank"]:self.items["shank"].append(str(i["shank"]).strip())
            if i["sub1No"] and i["sub1No"] not in self.items["sub"]:self.items["sub"].append(str(i["sub1No"]).strip())
            if i["sub2No"] and i["sub2No"] not in self.items["sub"]:self.items["sub"].append(str(i["sub2No"]).strip())
        else:
            [self.setText(*self.tableWidgetRCDict['tab{}Name'.format(idx)],x) for idx,x in enumerate(self.items["tab"],start=1)]
            [self.setText(*self.tableWidgetRCDict['seg{}No'.format(idx)],x) for idx,x in enumerate(self.items["segment"],start=1)]
            [self.setText(*self.tableWidgetRCDict['shank{}Name'.format(idx)],x) for idx,x in enumerate(self.items["shank"],start=1)]
            [self.setText(*self.tableWidgetRCDict['others{}Name'.format(idx)],x) for idx,x in enumerate(self.items["sub"],start=1)]
        self.groupSelected = True
    # ===========================================================================================    
    # 수량 수정시 호출
    def amountUpdate(self):

        self.amounts = {}
        for key, values in self.items.items():
            self.amounts[key] = {value: 0 for value in values}

        for idx,item in enumerate(self.selectedGroupItems):
            tabName = self.getTabName(item["groupName"],item["name"])
            try:
                tabCount = int(self.getText(*self.tableWidgetRCDict["tab{}Amount".format(idx+1)]))
                self.amounts["tab"][tabName] += tabCount
            except:
                ...
            else:
                def addAmounts(itemType, itemName, itemAmount):
                    try:
                        self.amounts[itemType][item[itemName]] += tabCount*int(item[itemAmount])
                    except ValueError:
                        ...
                    except:
                        print(f"err: update amounts: {itemType}-{itemName}-{itemAmount}")

                addAmounts("segment", "seg1No", "seg1Amount")
                addAmounts("segment", "seg2No", "seg2Amount")
                addAmounts("shank", "shank", "shankAmount")
                addAmounts("sub", "sub1No", "sub1Amount")
                addAmounts("sub", "sub2No", "sub2Amount")
        else:

            def writeAmount(itemType,posList,isSeg=True):
                for pos in [self.tableWidgetRCDict[x] for x in posList]:
                    r,c = pos
                    c = c+2 if isSeg else c+4
                    try:
                        amount = self.amounts[itemType][self.getText(*pos)]
                        if amount == 0:
                            self.setText(r,c,'')
                        else:
                            self.setText(r,c,str(amount))
                    except:
                        ...

            writeAmount("segment",["seg1No","seg2No"])
            writeAmount("shank",["shank1Name","shank2Name","shank3Name","shank4Name"],False)
            writeAmount("sub",["others1Name","others2Name","others3Name","others4Name"],False)

    # # 제조지시서 선택시 호출
    # def selectSporder(self,index):
    #     if index == 0: 
    #         self.clearAll()
    #         return
    #     sporderName = self.sporderNames[index-1]
    #     self.selectedSporder = self.ac.dm.getSingleItemDatas('sporder','mfgOrderNo',sporderName)
    #     self.setText(*self.tableWidgetRCDict['mfgOrderNo1'],self.selectedSporder['mfgOrderNo1'])
    #     self.setText(*self.tableWidgetRCDict['mfgOrderNo2'],self.selectedSporder['mfgOrderNo2'])
        
    #     self.loadSporder()

    # # 제조지시서 불러오기 (전체 내용)
    # def loadSporder(self):
    #     for k in self.selectedSporder.keys():
    #         try:
    #             #table widget
    #             if k in ['bondSelect', 'segNo1']:
    #                 self.getItem(*self.tableWidgetRCDict[k]).setCurrentText(self.selectedSporder[k])
    #             elif 'Check' in k:
    #                 self.getItem(*self.tableWidgetRCDict[k]).setChecked(True if self.selectedSporder[k] == '1' else False)
    #             else:
    #                 self.setText(*self.tableWidgetRCDict[k],self.selectedSporder[k])
    #         except: 
    #             #config table
    #             try:
    #                 self.setText(*self.configTableRCDict[k],self.selectedSporder[k], self.configTable)
    #             except:
    #                 #label
    #                 try:
    #                     text = self.labelDict[k].text().split(':')[0].strip()
    #                     self.labelDict[k].setText(f"{text} :  {str(self.selectedSporder[k])}")
    #                 except Exception as e:
    #                     print(e)

